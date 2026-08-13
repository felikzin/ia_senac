import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
import numpy as np
from sklearn.ensemble import IsolationForest


def aplicar_auditoria_na_planilha(caminho_arquivo):
    print("🔄 Processando planilha e aplicando auditoria de IA...")

    #carregar workbook
    wb = openpyxl.load_workbook(caminho_arquivo)
    if 'Financeiro' not in wb.sheetnames:
        raise ValueError("A aba 'Financeiro' não foi encontrada na planilha.")
    ws_fin = wb['Financeiro']

    #ler os dados da aba Financeiro
    dados = []
    for row in ws_fin.iter_rows(min_row=5, max_col=5, values_only=True):
        if row[1] is not None and str(row[1]).isdigit():
            dados.append({
                'NF': row[1],
                'Emissão': row[2].strftime('%d/%m/%Y') if hasattr(row[2], 'strftime') else row[2],
                'Vendedor': row[3],
                'Valor': float(row[4])
            })

    if not dados:
        print("⚠️ Nenhum dado de nota fiscal foi encontrado para analisar.")
        return
    df = pd.DataFrame(dados)

    #análise estatística e ML
    df['Media_Vendedor'] = df.groupby('Vendedor')['Valor'].transform('mean')
    df['Desvio_Vendedor'] = df.groupby('Vendedor')['Valor'].transform('std')
    df['Z_Score'] = (df['Valor'] - df['Media_Vendedor']) / df['Desvio_Vendedor']

    modelo_if = IsolationForest(contamination=0.03, random_state=42)
    df['Anomalia_ML'] = modelo_if.fit_predict(df[['Valor']])
    df['Status'] = np.where(
        (df['Anomalia_ML'] == -1) | (df['Z_Score'].abs() > 1.6),
        'ANOMALIA',
        'OK'
    )

    #criar ou resetar a aba Auditoria_IA
    if 'Auditoria_IA' in wb.sheetnames:
        del wb['Auditoria_IA']
    ws_aud = wb.create_sheet('Auditoria_IA')

    #exibir linhas de grade
    ws_aud.views.sheetView[0].showGridLines = True

    #estilização dos cabeçalhos
    ws_aud['B2'] = "RELATÓRIO DE AUDITORIA CONTINUA & MACHINE LEARNING"
    ws_aud['B2'].font = Font(name='Segoe UI', size=14, bold=True, color="1E293B")
    headers = ['NF', 'Emissão', 'Vendedor', 'Valor (R$)', 'Média Vendedor', 'Status Auditoria']
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=2):
        cell = ws_aud.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    #estilos de status
    fill_anomalia = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    font_anomalia = Font(name='Segoe UI', color="991B1B", bold=True)
    fill_ok = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    font_ok = Font(name='Segoe UI', color="166534")

    #escrever os dados na aba
    for r_idx, row in df.iterrows():
        current_row = r_idx + 5
        ws_aud.cell(row=current_row, column=2, value=row['NF']).alignment = Alignment(horizontal="center")
        ws_aud.cell(row=current_row, column=3, value=row['Emissão']).alignment = Alignment(horizontal="center")
        ws_aud.cell(row=current_row, column=4, value=row['Vendedor'])

        c_val = ws_aud.cell(row=current_row, column=5, value=row['Valor'])
        c_val.number_format = '_-"R$"\ * #,##0.00_-'
        c_med = ws_aud.cell(row=current_row, column=6, value=row['Media_Vendedor'])
        c_med.number_format = '_-"R$"\ * #,##0.00_-'

        texto_status = "!! ATENÇÃO" if row['Status'] == 'ANOMALIA' else "APROVADO"
        c_status = ws_aud.cell(row=current_row, column=7, value=texto_status)
        c_status.alignment = Alignment(horizontal="center")

        if row['Status'] == 'ANOMALIA':
            c_status.fill = fill_anomalia
            c_status.font = font_anomalia
        else:
            c_status.fill = fill_ok
            c_status.font = font_ok

    #salvar alterações
    wb.save(caminho_arquivo)
    print("✅ Sucesso! Nova aba 'Auditoria_IA' criada e atualizada na planilha.")


if __name__ == '__main__':
    caminho_planilha = r"C:\Users\augus\Downloads\teste.xlsx"
    aplicar_auditoria_na_planilha(caminho_planilha)
